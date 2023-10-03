from tempfile import NamedTemporaryFile
from openpyxl import Workbook
from openpyxl.styles import PatternFill,Font, Alignment, Border, Side
from sqlalchemy import create_engine, text


def create_csv(id):
    engine = create_engine("sqlite:///app.db") # database connection



    deliveries_query="SELECT Deliveries.deliveries_id as DeliveriesID, Deliveries.point_id as PointID, Deliveries.deliveries_date as DateDeliver, Deliveries.deliveries_price as Price, Product.product ,Deliveries.quantity as Quаntity FROM Deliveries LEFT JOIN Product ON Deliveries.product_id=Product.product_id_warehouse WHERE point_id = '{}'".format(id) #запросы к каждой из таблиц (вся информация)
    expenses_query = "SELECT expenses_id as expensesID, rent, services, salary, date as expenseDate FROM Expenses WHERE point_id = '{}'".format(id)
    product_query = "SELECT product_id_warehouse as ID, product, price  FROM Product"
    receipt_query = "SELECT receipt_number as number, Product.product, quantity, Receipt.price, date  FROM Receipt LEFT JOIN Product ON Receipt.product_id=Product.product_id_warehouse WHERE point_id = '{}'".format(id)

    #Отчет
    report_query = "SELECT " \
                   "(SELECT SUM(price) FROM Receipt WHERE  Receipt.point_id = '{}') as Income,"\
                   "(SELECT SUM(deliveries_price) FROM Deliveries WHERE  Deliveries.point_id = '{}') as Deliveries, " \
                   "(SELECT SUM(salary + rent + services) FROM Expenses WHERE Expenses.point_id = '{}') as Expenses".format(id, id, id)  # Запрос для отчета SUM(price) Receipt

    with engine.connect() as connection:
        deliv_result = connection.execute(text(deliveries_query))
        exp_result = connection.execute(text(expenses_query))
        prod_result = connection.execute(text(product_query))
        receipt_result = connection.execute(text(receipt_query))

        report_result = connection.execute(text(report_query)) #Отчет

    wb = Workbook()

    ws1=wb.active # work with default worksheet
    ws1.title = "Deliveries"
    l1=[r for r in deliv_result.keys()] # List of column headers
    ws1.append(l1) # adding column headers at first row

    ws2 = wb.create_sheet("Expenses")
    l2 = [r for r in exp_result.keys()] # List of column headers
    ws2.append(l2) # adding column headers at first row

    ws3 = wb.create_sheet("Product")
    l3 = [r for r in prod_result.keys()] # List of column headers
    ws3.append(l3) # adding column headers at first row

    ws4 = wb.create_sheet("Receipt")
    l4 = [r for r in receipt_result.keys()]  # List of column headers
    ws4.append(l4)  # adding column headers at first row

    ws5 = wb.create_sheet("Report")
    l5 = [r for r in report_result.keys()]
    ws5.append(l5)

    ws_list = [ws1, ws2, ws3, ws4, ws5]


    my_font=Font(size=12,bold=True) # font styles
    # my_fill=PatternFill(fill_type='solid',start_color='FFFF00') #Background color
    my_border = Border(left=Side(border_style='dotted',color='000000'),
                       right=Side(border_style='dotted',color='000000'),
                       top=Side(border_style='dotted', color='000000'),
                       bottom=Side(border_style='dotted', color='000000'))


    #styles
    for i in ws_list:
        for column in i.columns:
            max_length = 0
            column_letter = column[0].column_letter
            i.border = my_border
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            i.column_dimensions[column_letter].width = adjusted_width


    for cell in ws_list:
        for i in cell["1:1"]:
            i.font = my_font
            i.alignment = Alignment(horizontal="center")
            i.border = my_border

    r, c = 2, 0  # row=2 and column=0

    #info
    result_list = [deliv_result, exp_result, prod_result, receipt_result, report_result]
    ws_index = 0
    for i in result_list:
        for row_data in i:
            d = [r for r in row_data]
            ws_list[ws_index].append(d)
        ws_index += 1



    # wb.save("file_path") # сохранить файл локально

####################Save file as Stream ####################################
    with NamedTemporaryFile(delete=False) as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        create_csv.stream = tmp.read()
        # print(tmp.name)
        create_csv.csv_path = tmp.name


create_csv(id)

